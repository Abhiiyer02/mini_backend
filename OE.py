import openpyxl as op

class Student:
    def __init__(self, name, sem, section, USN, email, CGPA, branch, choices):
        self.name = name
        self.USN = USN
        self.CGPA = CGPA
        self.branch = branch
        self.ochoices=choices
        self.alloc = ""
        #self.phone=pno
        self.email=email
        self.section=section
        self.sem=sem
        self.choices=[]
        for i in choices :
            if i not in self.choices :
                self.choices.append(i) 

    def __str__(self):
        return(f"Name : {self.name} , USN : {self.USN} , CGPA : {self.CGPA} , branch : {self.branch} , choices : {self.choices} , allocated elective : {self.alloc}")

    def tuplestr(self):
        return(f"{self.name},{self.sem},{self.section},{self.USN},{self.email},{self.CGPA},{self.branch},{self.alloc},{(',').join(self.ochoices)}") 

    def tuplestr2(self):
        return(f"{self.name},{self.USN},{self.CGPA},{self.branch}")

class Elective:
    def __init__(self,courseid,ename,MaxCap=60):
        self.Name = ename
        self.courseID = courseid
        #self.Branch = branch
        self.MaxCap = MaxCap
        self.NoStuds = 0
        self.MinCgpa = float('inf')
        self.same = []

    def __str__(self):
        return(f"Name : {self.Name} , ID : {self.courseID}  , No. of students: {self.NoStuds}, MaxCap : {self.MaxCap}, Min CGPA : {self.MinCgpa}")
    
    def codeName(self):
        return(f"{self.courseID}: {self.Name}")
    
    def countStuds(self):
        return(f"{self.courseID},{self.Name},{self.NoStuds},{self.MinCgpa}")

def createElectives():
    courseCode= ""
    for i in range(2,courses.max_row+1):
        for j in range (1,courses.max_column+1):
            val = courses.cell(i,j).value
            if j ==1 :
                courseCode = val
            elif j ==2 :
                ename = val
            #elif j ==3 :
            #    branch = val    
        e = Elective(courseCode,ename)#,branch)
        Course[courseCode]=e
        
def createStud():
    for i in range(2,studs.max_row+1):
        choice=[]
        for j in range(2,19):
            val=studs.cell(i,j).value
            #print(val)
            if j ==2:
                email=val
            elif j ==3:
                usn = val
            elif j ==4:
                sname = val
            elif j ==5:
                #print(i,j)
                cgpa = float(val)
            elif j==6:
                sem=val
            elif j ==7 :
                sec=val      
            elif j ==8 :
                sbranch = val
            else:
                if val != None:
                    val=val.split(': ')[0]
                    choice.append(val)#val[0:-1]
        s=Student(sname,sem,sec,usn,email,cgpa,sbranch,choice)
        Studs.append(s)

def OriginalForm() :
    for s in Studs:
        if s.alloc != '' :
            s.alloc=Course[s.alloc].codeName()
        for j in s.ochoices:
            if j  in Course.keys():
                s.ochoices=list(map(lambda x: x.replace(j,Course[j].codeName()),s.ochoices))

def RawResult(Studs):
    newWorkbook = op.load_workbook("Allotment2.xlsx")
    sheet=newWorkbook.active
    l=[]
    t1=("Student Name","Semester","Section","USN","Email ID","CGPA","Branch","Allotted Course","Choice 1","Choice 2","Choice 3","Choice 4","Choice 5","Choice 6","Choice 7","Choice 8","Choice 9","Choice 10")
    l.append(t1)
    for i in Studs:
        v=i.tuplestr().split(',')
        t=tuple(v)
        l.append(t)
    for j in l:
        sheet.append(j)
        
    newWorkbook.save(filepath)

def writeSummary():
    newWorkbook = op.load_workbook("Allotment2.xlsx")
    newWorkbook.active=newWorkbook["Summary"]
    t1=("Course Code","Course Name","No of students","Minimum CGPA/Cut off")
    l=[]
    l.append(t1)
    for j in Course.keys() :
            u=Course[j].countStuds().split(',')
            t1=tuple(u)
            l.append(t1)
    for j in l:
        newWorkbook.active.append(j)

    newWorkbook.save(filepath)        

def writeCourses(sheet_names,Stud):
    newWorkbook = op.load_workbook("Allotment2.xlsx")
    for i in sheet_names:
        newWorkbook.active = newWorkbook[i]
        l=[]
        l.append(tuple([Course[i].codeName()]))
        t1=("Student Name","USN","CGPA","Branch")
        l.append(t1)
        for j in Stud :
            if((j.alloc.split(': ')[0])==i) :
                u=j.tuplestr2().split(',')
                t1=tuple(u)
                l.append(t1)
        for j in l:
            newWorkbook.active.append(j)
    newWorkbook.save(filepath)
                                   
def AllotCourse(Stud,i):
    for j in Stud.choices:
        if j in Course.keys():
            if Course[j].NoStuds < Course[j].MaxCap :
                Stud.alloc=j
                Course[j].NoStuds+=1
                if Stud.CGPA < Course[j].MinCgpa :
                    Course[j].MinCgpa=Stud.CGPA
                    Course[j].same=[]
                    Course[j].same.append(i)
                elif Stud.CGPA == Course[j].MinCgpa :
                    Course[j].same.append(i)
                return -1
            
            elif Course[j].NoStuds < (Course[j].MaxCap + buffer) :
                if Stud.CGPA == Course[j].MinCgpa :
                    Stud.alloc=j
                    Course[j].NoStuds+=1
                    Course[j].same.append(i)
                    return -1

            elif  Course[j].NoStuds == (Course[j].MaxCap + buffer) :
                if Stud.CGPA == Course[j].MinCgpa :
                    prev=Course[j].same[0]
                    for k in range(1,len(Course[j].same)) :
                        if Studs[Course[j].same[k]].choices.index(j) > Studs[prev].choices.index(j):
                            prev=Course[j].same[k]                  
                    if Stud.choices.index(j) < Studs[prev].choices.index(j) :
                        Stud.alloc=j
                        Course[j].same.remove(prev)
                        Course[j].same.append(i)
                        return prev
                
    return -1        

def StartAllotment() :
    for i in range(len(Studs)):
        reallot= -1
        reallot=AllotCourse(Studs[i],i)
        while reallot > 0 :
            Studs[reallot].alloc=""
            reallot=AllotCourse(Studs[reallot],reallot)

def  CheckCourse30() :
    sheet_names=list(Course.keys())
    newAllot=False
    leastCourse=[]
    for i in sheet_names :
        if Course[i].NoStuds < 30 :
            leastCourse.append(i)
            newAllot=True
    if len(leastCourse) > 0 :        
        Coursedel=leastCourse[0]
        for j in leastCourse :
            if Course[j].NoStuds < Course[Coursedel].NoStuds :
                Coursedel=j  
        del Course[Coursedel]  
        blockedCourses.append(Coursedel)    

    if newAllot == True :
        for s in Studs :
            s.alloc=""
            while Coursedel in s.choices:
                s.choices.remove(Coursedel)
        sheet_names.remove(Coursedel)
        for i in sheet_names:
            Course[i].NoStuds=0
            Course[i].same=[]
            Course[i].MinCgpa=float('inf')

    return newAllot       

if __name__ == "__main__":
    wb=op.load_workbook("RawDataOE1.xlsx")    
    filepath = r"D:/ABHIJITH IYER R/jce/notes , textbooks etc/sem 5/open elective allotment project/Allotment2.xlsx" 
    #r"C:/Users/Admin/Downloads/OpenElectives/Allotment3.xlsx"
    #
    # filepath = "d:/CSE/5TH SEMESTER/Open Elective Allotment/Allotment3.xlsx" 
    #  r"D:/codes/open elective/Allotment3.xlsx"

    studs=wb['OE2 RawData']
    courses = wb['OE2 Courses']
    Studs=[]
    Course = {}
    blockedCourses=[]
    buffer=0

    createElectives()
    createStud()
                                                          
    Studs.sort(key = lambda Stud: Stud.CGPA,reverse=True)

    StartAllotment()
    Studs=sorted(Studs,key = lambda Stud: Stud.USN)                
                  
    res = op.Workbook()
    sheet=res.active
    sheet.title="Allotment"
    sheet_names=list(Course.keys())
    ws=[]
    
    print(blockedCourses)
    wsx=res.create_sheet(index=1,title="Summary")
    ws.append(wsx)
    for i in range(len(sheet_names)):
        wsx=res.create_sheet(index=(i+2),title=sheet_names[i])
        ws.append(wsx)
    for i in sheet_names:
        print(Course[i])
    res.save(filepath)
    
    OriginalForm()
    RawResult(Studs)
    writeSummary()
    writeCourses(sheet_names,Studs)